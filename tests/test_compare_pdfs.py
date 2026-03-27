# Copyright (c) 2026 MediSage Co,Ltd.
# SPDX-License-Identifier: MIT
"""Unit tests for compare_pdfs."""

from __future__ import annotations

import cv2
import numpy as np
import pytest

from compare_pdfs import (
    DEFAULT_CONFIG,
    DiffRegion,
    PageResult,
    _merge_report_rows,
    _sanitize_sheet_name,
    _scale_image,
    build_parser,
    create_overlay,
    decide_status,
    deep_merge,
    detect_diff,
    load_config,
    main,
    normalize_size,
    pair_pdfs,
    write_csv,
    write_xlsx,
)


# ---------------------------------------------------------------------------
# deep_merge
# ---------------------------------------------------------------------------
class TestDeepMerge:
    def test_flat(self):
        assert deep_merge({"a": 1}, {"b": 2}) == {"a": 1, "b": 2}

    def test_override(self):
        assert deep_merge({"a": 1}, {"a": 2}) == {"a": 2}

    def test_nested(self):
        base = {"x": {"a": 1, "b": 2}}
        update = {"x": {"b": 3, "c": 4}}
        assert deep_merge(base, update) == {"x": {"a": 1, "b": 3, "c": 4}}

    def test_empty_update(self):
        base = {"a": 1}
        assert deep_merge(base, {}) == {"a": 1}


# ---------------------------------------------------------------------------
# load_config
# ---------------------------------------------------------------------------
class TestLoadConfig:
    def test_default(self):
        config = load_config(None)
        assert config["dpi"] == 300
        assert config["status_thresholds"]["ng_min_regions"] == 8

    def test_override_from_file(self, tmp_path):
        cfg = tmp_path / "config.yaml"
        cfg.write_text("dpi: 150\n", encoding="utf-8")
        config = load_config(cfg)
        assert config["dpi"] == 150
        # Other defaults preserved
        assert config["pixel_threshold"] == 10

    def test_nested_override(self, tmp_path):
        cfg = tmp_path / "config.yaml"
        cfg.write_text("status_thresholds:\n  ng_min_regions: 20\n", encoding="utf-8")
        config = load_config(cfg)
        assert config["status_thresholds"]["ng_min_regions"] == 20
        assert config["status_thresholds"]["ng_min_area"] == 3000


# ---------------------------------------------------------------------------
# normalize_size
# ---------------------------------------------------------------------------
class TestNormalizeSize:
    def test_same_size(self):
        img = np.zeros((100, 200, 3), dtype=np.uint8)
        a, b = normalize_size(img, img, crop_to_common_size=True)
        assert a.shape == b.shape == (100, 200, 3)

    def test_crop(self):
        img1 = np.zeros((100, 200, 3), dtype=np.uint8)
        img2 = np.zeros((80, 150, 3), dtype=np.uint8)
        a, b = normalize_size(img1, img2, crop_to_common_size=True)
        assert a.shape[:2] == (80, 150)
        assert b.shape[:2] == (80, 150)

    def test_no_crop_raises(self):
        img1 = np.zeros((100, 200, 3), dtype=np.uint8)
        img2 = np.zeros((80, 150, 3), dtype=np.uint8)
        with pytest.raises(ValueError, match="image size mismatch"):
            normalize_size(img1, img2, crop_to_common_size=False)


# ---------------------------------------------------------------------------
# detect_diff
# ---------------------------------------------------------------------------
class TestDetectDiff:
    def test_identical_images(self):
        img = np.full((100, 100, 3), 128, dtype=np.uint8)
        _, regions, area, ratio = detect_diff(img, img, pixel_threshold=10, min_region_area=20, morph_kernel=3)
        assert regions == []
        assert area == 0
        assert ratio == 0.0

    def test_different_images(self):
        img1 = np.full((100, 100, 3), 0, dtype=np.uint8)
        img2 = np.full((100, 100, 3), 0, dtype=np.uint8)
        # Draw a white rectangle on img2
        cv2.rectangle(img2, (20, 20), (60, 60), (255, 255, 255), cv2.FILLED)
        _, regions, area, ratio = detect_diff(img1, img2, pixel_threshold=10, min_region_area=20, morph_kernel=1)
        assert len(regions) > 0
        assert area > 0
        assert ratio > 0.0

    def test_small_diff_filtered_by_min_area(self):
        img1 = np.full((100, 100, 3), 0, dtype=np.uint8)
        img2 = img1.copy()
        # Single pixel change
        img2[50, 50] = [255, 255, 255]
        _, regions, _, _ = detect_diff(img1, img2, pixel_threshold=10, min_region_area=20, morph_kernel=1)
        assert len(regions) == 0


# ---------------------------------------------------------------------------
# create_overlay
# ---------------------------------------------------------------------------
class TestCreateOverlay:
    def test_output_shape(self):
        img_old = np.full((100, 100, 3), 200, dtype=np.uint8)
        img_new = np.full((100, 100, 3), 100, dtype=np.uint8)
        regions = [DiffRegion(x=10, y=10, w=30, h=30, area=900)]
        overlay = create_overlay(img_old, img_new, regions)
        assert overlay.shape == (100, 100, 3)

    def test_no_regions(self):
        img = np.full((50, 50, 3), 128, dtype=np.uint8)
        overlay = create_overlay(img, img, [])
        assert overlay.shape == (50, 50, 3)


# ---------------------------------------------------------------------------
# decide_status
# ---------------------------------------------------------------------------
class TestDecideStatus:
    thresholds = DEFAULT_CONFIG["status_thresholds"]

    def test_ok(self):
        status, reason = decide_status(0, 0, 0.0, self.thresholds)
        assert status == "OK"
        assert reason == "no_diff"

    def test_review_by_regions(self):
        status, reason = decide_status(1, 0, 0.0, self.thresholds)
        assert status == "REVIEW"
        assert reason == "threshold_review"

    def test_review_by_area(self):
        status, reason = decide_status(0, 1, 0.0, self.thresholds)
        assert status == "REVIEW"
        assert reason == "threshold_review"

    def test_ng_by_regions(self):
        status, reason = decide_status(8, 0, 0.0, self.thresholds)
        assert status == "NG"
        assert reason == "threshold_ng"

    def test_ng_by_area(self):
        status, reason = decide_status(0, 3000, 0.0, self.thresholds)
        assert status == "NG"
        assert reason == "threshold_ng"

    def test_ng_by_ratio(self):
        status, reason = decide_status(0, 0, 0.0015, self.thresholds)
        assert status == "NG"
        assert reason == "threshold_ng"


# ---------------------------------------------------------------------------
# pair_pdfs
# ---------------------------------------------------------------------------
class TestPairPdfs:
    def test_basic_pairing(self, tmp_path):
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()
        (old_dir / "a.pdf").touch()
        (old_dir / "b.pdf").touch()
        (new_dir / "a.pdf").touch()
        (new_dir / "c.pdf").touch()

        paired, warnings, old_only, new_only = pair_pdfs(old_dir, new_dir)
        assert len(paired) == 1
        assert paired[0][0].name == "a.pdf"
        assert old_only == ["b.pdf"]
        assert new_only == ["c.pdf"]
        assert len(warnings) == 2

    def test_empty_dirs(self, tmp_path):
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()
        paired, warnings, old_only, new_only = pair_pdfs(old_dir, new_dir)
        assert paired == []
        assert old_only == []
        assert new_only == []


# ---------------------------------------------------------------------------
# _scale_image
# ---------------------------------------------------------------------------
class TestScaleImage:
    def test_scale_down(self):
        img = np.zeros((800, 600, 3), dtype=np.uint8)
        result = _scale_image(img, 300)
        assert result.shape[1] == 300
        assert result.shape[0] == 400  # 800 * (300/600)

    def test_scale_up(self):
        img = np.zeros((100, 50, 3), dtype=np.uint8)
        result = _scale_image(img, 200)
        assert result.shape[1] == 200
        assert result.shape[0] == 400  # 100 * (200/50)


# ---------------------------------------------------------------------------
# _sanitize_sheet_name
# ---------------------------------------------------------------------------
class TestSanitizeSheetName:
    def test_clean_name(self):
        assert _sanitize_sheet_name("hello") == "hello"

    def test_special_chars(self):
        assert _sanitize_sheet_name("a[b]c*d?e") == "a_b_c_d_e"

    def test_truncate(self):
        name = "a" * 50
        assert len(_sanitize_sheet_name(name)) == 31

    def test_custom_max_len(self):
        assert len(_sanitize_sheet_name("abcdefghij", max_len=5)) == 5


# ---------------------------------------------------------------------------
# _merge_report_rows
# ---------------------------------------------------------------------------
class TestMergeReportRows:
    def _make_result(self, file: str, page: int, status: str = "OK") -> PageResult:
        return PageResult(
            file=file,
            page=page,
            old_width=100,
            old_height=100,
            new_width=100,
            new_height=100,
            status=status,
            diff_regions=0,
            diff_area=0,
            diff_ratio=0.0,
            review_reason="no_diff",
            diff_image=None,
            blocks_old=0,
            blocks_new=0,
            block_count_delta=0,
            regions=[],
        )

    def test_sorted_by_filename(self):
        results = [self._make_result("c.pdf", 1), self._make_result("a.pdf", 1)]
        rows = _merge_report_rows(results, [], [])
        assert rows[0][0] == "a.pdf"
        assert rows[1][0] == "c.pdf"

    def test_old_only_new_only_interleaved(self):
        results = [self._make_result("b.pdf", 1)]
        rows = _merge_report_rows(results, ["a.pdf"], ["c.pdf"])
        assert [r[0] for r in rows] == ["a.pdf", "b.pdf", "c.pdf"]
        assert rows[0][6] == "OLD_ONLY"
        assert rows[1][6] == "OK"
        assert rows[2][6] == "NEW_ONLY"

    def test_multi_page_ordering(self):
        results = [
            self._make_result("a.pdf", 3),
            self._make_result("a.pdf", 1),
            self._make_result("a.pdf", 2),
        ]
        rows = _merge_report_rows(results, [], [])
        assert [r[1] for r in rows] == [1, 2, 3]

    def test_empty(self):
        assert _merge_report_rows([], [], []) == []


# ---------------------------------------------------------------------------
# write_csv
# ---------------------------------------------------------------------------
class TestWriteCsv:
    def _make_result(self, file: str, page: int) -> PageResult:
        return PageResult(
            file=file,
            page=page,
            old_width=100,
            old_height=100,
            new_width=100,
            new_height=100,
            status="OK",
            diff_regions=0,
            diff_area=0,
            diff_ratio=0.0,
            review_reason="no_diff",
            diff_image=None,
            blocks_old=1,
            blocks_new=1,
            block_count_delta=0,
            regions=[],
        )

    def test_output_sorted(self, tmp_path):
        path = tmp_path / "report.csv"
        results = [self._make_result("c.pdf", 1), self._make_result("a.pdf", 1)]
        write_csv(path, results, ["b.pdf"], [])
        lines = path.read_text(encoding="utf-8-sig").strip().split("\n")
        assert len(lines) == 4  # header + 3 rows
        assert lines[1].startswith("a.pdf")
        assert lines[2].startswith("b.pdf")
        assert lines[3].startswith("c.pdf")


# ---------------------------------------------------------------------------
# write_xlsx
# ---------------------------------------------------------------------------
class TestWriteXlsx:
    def _make_result(self, file: str, page: int, status: str = "OK") -> PageResult:
        return PageResult(
            file=file,
            page=page,
            old_width=100,
            old_height=100,
            new_width=100,
            new_height=100,
            status=status,
            diff_regions=0,
            diff_area=0,
            diff_ratio=0.0,
            review_reason="no_diff",
            diff_image=None,
            blocks_old=1,
            blocks_new=1,
            block_count_delta=0,
            regions=[],
        )

    def test_report_sheet_sorted(self, tmp_path):
        path = tmp_path / "report.xlsx"
        results = [self._make_result("c.pdf", 1), self._make_result("a.pdf", 1)]
        write_xlsx(path, results, ["b.pdf"], [], {})
        wb = __import__("openpyxl").load_workbook(path)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "a.pdf"
        assert ws.cell(row=3, column=1).value == "b.pdf"
        assert ws.cell(row=4, column=1).value == "c.pdf"

    def test_no_image_sheets_when_no_images(self, tmp_path):
        path = tmp_path / "report.xlsx"
        results = [self._make_result("a.pdf", 1, "NG")]
        write_xlsx(path, results, [], [], {})
        wb = __import__("openpyxl").load_workbook(path)
        assert len(wb.sheetnames) == 1


# ---------------------------------------------------------------------------
# Fixture: tiny test PDFs
# ---------------------------------------------------------------------------
def _make_minimal_pdf(text: str, rect: tuple | None = None) -> bytes:
    """Generate a minimal single-page 200x200 PDF with text and optional rect."""
    objects = [
        b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj",
        b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj",
        b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 200 200]"
        b" /Contents 5 0 R /Resources << /Font << /F1 4 0 R >> >> >>\nendobj",
        b"4 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj",
    ]
    stream = f"BT /F1 14 Tf 10 150 Td ({text}) Tj ET"
    if rect:
        x1, y1, x2, y2 = rect
        stream += f" {x1} {y1} {x2 - x1} {y2 - y1} re S"
    sb = stream.encode()
    objects.append(f"5 0 obj\n<< /Length {len(sb)} >>\nstream\n".encode() + sb + b"\nendstream\nendobj")

    header = b"%PDF-1.4\n"
    body = header
    offsets = []
    for obj in objects:
        offsets.append(len(body))
        body += obj + b"\n"
    xref_pos = len(body)
    xref = f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n"
    for off in offsets:
        xref += f"{off:010d} 00000 n \n"
    xref += f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF\n"
    return body + xref.encode()


@pytest.fixture()
def pdf_dirs(tmp_path):
    """Create old/new dirs with small test PDFs (identical + different)."""
    old_dir = tmp_path / "old"
    new_dir = tmp_path / "new"
    old_dir.mkdir()
    new_dir.mkdir()

    # same.pdf — identical in both dirs
    same_pdf = _make_minimal_pdf("same")
    (old_dir / "same.pdf").write_bytes(same_pdf)
    (new_dir / "same.pdf").write_bytes(same_pdf)

    # diff.pdf — different between dirs
    (old_dir / "diff.pdf").write_bytes(_make_minimal_pdf("OLD", (50, 50, 150, 120)))
    (new_dir / "diff.pdf").write_bytes(_make_minimal_pdf("NEW", (60, 60, 160, 130)))

    # only_old.pdf — only in old
    (old_dir / "only_old.pdf").write_bytes(_make_minimal_pdf("old_only"))

    return old_dir, new_dir


# ---------------------------------------------------------------------------
# build_parser
# ---------------------------------------------------------------------------
class TestBuildParser:
    def test_defaults(self):
        parser = build_parser()
        args = parser.parse_args(["--old-dir", "a", "--new-dir", "b"])
        assert args.format == "xlsx"
        assert args.without_diff_image is False

    def test_format_csv(self):
        parser = build_parser()
        args = parser.parse_args(["--old-dir", "a", "--new-dir", "b", "--format", "csv"])
        assert args.format == "csv"

    def test_without_diff_image(self):
        parser = build_parser()
        args = parser.parse_args(["--old-dir", "a", "--new-dir", "b", "--without-diff-image"])
        assert args.without_diff_image is True


# ---------------------------------------------------------------------------
# main (E2E)
# ---------------------------------------------------------------------------
class TestMain:
    def test_xlsx_default(self, pdf_dirs, tmp_path):
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        rc = main(["--old-dir", str(old_dir), "--new-dir", str(new_dir), "--output-dir", str(out)])
        assert rc in (0, 1)  # 1 because of old_only warning
        assert (out / "report.xlsx").exists()
        assert (out / "diff_images").is_dir()
        # temp images cleaned up
        assert not (out / ".temp_images").exists()

    def test_csv_format(self, pdf_dirs, tmp_path):
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        main(["--old-dir", str(old_dir), "--new-dir", str(new_dir), "--output-dir", str(out), "--format", "csv"])
        assert (out / "report.csv").exists()
        assert not (out / "report.xlsx").exists()

    def test_json_format(self, pdf_dirs, tmp_path):
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        main(["--old-dir", str(old_dir), "--new-dir", str(new_dir), "--output-dir", str(out), "--format", "json"])
        assert (out / "report.json").exists()

    def test_without_diff_image(self, pdf_dirs, tmp_path):
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        main(
            [
                "--old-dir",
                str(old_dir),
                "--new-dir",
                str(new_dir),
                "--output-dir",
                str(out),
                "--without-diff-image",
            ]
        )
        assert (out / "report.xlsx").exists()
        assert not (out / "diff_images").exists()

    def test_xlsx_report_sorted(self, pdf_dirs, tmp_path):
        """Verify rows in xlsx are sorted by filename."""
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        main(["--old-dir", str(old_dir), "--new-dir", str(new_dir), "--output-dir", str(out)])
        import openpyxl

        wb = openpyxl.load_workbook(out / "report.xlsx")
        ws = wb.active
        filenames = []
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=1).value
            if val:
                filenames.append(val)
        assert filenames == sorted(filenames)

    def test_csv_report_sorted(self, pdf_dirs, tmp_path):
        """Verify rows in CSV are sorted by filename."""
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        main(["--old-dir", str(old_dir), "--new-dir", str(new_dir), "--output-dir", str(out), "--format", "csv"])
        import csv as csv_mod

        with (out / "report.csv").open(encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            next(reader)  # skip header
            filenames = [row[0] for row in reader]
        assert filenames == sorted(filenames)

    def test_limit(self, pdf_dirs, tmp_path):
        old_dir, new_dir = pdf_dirs
        out = tmp_path / "out"
        main(
            [
                "--old-dir",
                str(old_dir),
                "--new-dir",
                str(new_dir),
                "--output-dir",
                str(out),
                "--format",
                "csv",
                "--limit",
                "1",
            ]
        )
        import csv as csv_mod

        with (out / "report.csv").open(encoding="utf-8-sig") as f:
            reader = csv_mod.reader(f)
            next(reader)  # skip header
            rows = list(reader)
        # 1 paired file processed + 1 old_only = at most 2 data rows
        paired_rows = [r for r in rows if r[6] not in ("OLD_ONLY", "NEW_ONLY")]
        assert len(paired_rows) <= 1

    def test_no_matching_files(self, tmp_path):
        old_dir = tmp_path / "old"
        new_dir = tmp_path / "new"
        old_dir.mkdir()
        new_dir.mkdir()
        (old_dir / "a.pdf").touch()
        (new_dir / "b.pdf").touch()
        rc = main(["--old-dir", str(old_dir), "--new-dir", str(new_dir), "--output-dir", str(tmp_path / "out")])
        assert rc == 2
