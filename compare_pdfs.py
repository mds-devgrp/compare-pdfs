#!/usr/bin/env python3
# Copyright (c) 2026 MediSage Co,Ltd.
# SPDX-License-Identifier: MIT
"""Batch PDF layout diff tool.

Features
- Pair PDFs by filename from old/new directories
- Render pages with pypdfium2 at a fixed DPI
- Optional image alignment (ECC) to suppress global shift noise
- Detect very small layout differences with tunable thresholding
- Output annotated diff images, CSV, and JSON summary
- Optional text block extraction summary via PyMuPDF
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import cv2
import numpy as np
import openpyxl
import pypdfium2 as pdfium
import yaml
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

__version__ = "dev"


@dataclass
class DiffRegion:
    x: int
    y: int
    w: int
    h: int
    area: int


@dataclass
class PageResult:
    file: str
    page: int
    old_width: int
    old_height: int
    new_width: int
    new_height: int
    status: str
    diff_regions: int
    diff_area: int
    diff_ratio: float
    review_reason: str
    diff_image: Optional[str]
    blocks_old: int
    blocks_new: int
    block_count_delta: int
    regions: List[DiffRegion]


DEFAULT_CONFIG = {
    "dpi": 300,
    "pixel_threshold": 10,
    "min_region_area": 20,
    "morph_kernel": 3,
    "align_images": True,
    "ecc_iterations": 200,
    "ecc_eps": 1e-5,
    "crop_to_common_size": True,
    "page_count_policy": "error",
    "status_thresholds": {
        "review_min_regions": 1,
        "review_min_area": 1,
        "ng_min_regions": 8,
        "ng_min_area": 3000,
        "ng_min_ratio": 0.0015,
    },
}


class ConfigError(Exception):
    pass


def load_config(path: Optional[Path]) -> dict:
    config = json.loads(json.dumps(DEFAULT_CONFIG))
    if path is None:
        return config
    with path.open("r", encoding="utf-8") as f:
        user_config = yaml.safe_load(f) or {}
    if not isinstance(user_config, dict):
        raise ConfigError("config.yaml must be a mapping")
    return deep_merge(config, user_config)


def deep_merge(base: dict, update: dict) -> dict:
    result = dict(base)
    for key, value in update.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def render_pdf_pages(pdf_path: Path, dpi: int) -> List[np.ndarray]:
    scale = dpi / 72.0
    pdf = pdfium.PdfDocument(str(pdf_path))
    pages: List[np.ndarray] = []
    for i in range(len(pdf)):
        page = pdf[i]
        bitmap = page.render(scale=scale)
        arr = bitmap.to_numpy()
        if arr.ndim == 3 and arr.shape[2] == 4:
            bgr = cv2.cvtColor(arr, cv2.COLOR_BGRA2BGR)
        elif arr.ndim == 3 and arr.shape[2] == 3:
            bgr = arr
        else:
            bgr = cv2.cvtColor(arr, cv2.COLOR_GRAY2BGR)
        pages.append(bgr)
    return pages


def normalize_size(img1: np.ndarray, img2: np.ndarray, crop_to_common_size: bool) -> Tuple[np.ndarray, np.ndarray]:
    if img1.shape[:2] == img2.shape[:2]:
        return img1, img2
    if not crop_to_common_size:
        raise ValueError(f"image size mismatch: {img1.shape[:2]} vs {img2.shape[:2]}")
    h = min(img1.shape[0], img2.shape[0])
    w = min(img1.shape[1], img2.shape[1])
    return img1[:h, :w].copy(), img2[:h, :w].copy()


def try_align_images(base_img: np.ndarray, target_img: np.ndarray, iterations: int, eps: float) -> np.ndarray:
    gray_base = cv2.cvtColor(base_img, cv2.COLOR_BGR2GRAY)
    gray_target = cv2.cvtColor(target_img, cv2.COLOR_BGR2GRAY)
    warp_matrix = np.eye(2, 3, dtype=np.float32)
    criteria = (cv2.TERM_CRITERIA_EPS | cv2.TERM_CRITERIA_COUNT, iterations, eps)
    try:
        _, warp_matrix = cv2.findTransformECC(
            gray_base,
            gray_target,
            warp_matrix,
            cv2.MOTION_TRANSLATION,
            criteria,
        )
        aligned = cv2.warpAffine(
            target_img,
            warp_matrix,
            (base_img.shape[1], base_img.shape[0]),
            flags=cv2.INTER_LINEAR + cv2.WARP_INVERSE_MAP,
            borderMode=cv2.BORDER_CONSTANT,
            borderValue=(255, 255, 255),
        )
        return aligned
    except cv2.error:
        return target_img


def detect_diff(
    img1: np.ndarray,
    img2: np.ndarray,
    pixel_threshold: int,
    min_region_area: int,
    morph_kernel: int,
) -> Tuple[np.ndarray, List[DiffRegion], int, float]:
    diff = cv2.absdiff(img1, img2)
    gray = cv2.cvtColor(diff, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, pixel_threshold, 255, cv2.THRESH_BINARY)

    if morph_kernel > 1:
        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (morph_kernel, morph_kernel))
        thresh = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, kernel)
        thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)

    contours, _ = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    regions: List[DiffRegion] = []
    total_area = 0
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = int(cv2.contourArea(contour))
        if area < min_region_area:
            continue
        regions.append(DiffRegion(x=x, y=y, w=w, h=h, area=area))
        total_area += area

    total_pixels = thresh.shape[0] * thresh.shape[1]
    ratio = float(total_area / total_pixels) if total_pixels else 0.0
    return thresh, sorted(regions, key=lambda r: (r.y, r.x)), total_area, ratio


def imwrite_unicode(path: Path, img: np.ndarray) -> bool:
    """cv2.imwrite that supports non-ASCII (e.g. Japanese) file paths."""
    result, buf = cv2.imencode(path.suffix, img)
    if not result:
        return False
    path.write_bytes(buf.tobytes())
    return True


def create_overlay(img_old: np.ndarray, img_new: np.ndarray, regions: Sequence[DiffRegion]) -> np.ndarray:
    """Create a blended overlay image highlighting diff regions in red."""
    # Blend old (blue tint) and new (green tint) at 50% each
    blended = cv2.addWeighted(img_old, 0.5, img_new, 0.5, 0)

    # Build a mask from diff regions and tint those areas red
    mask = np.zeros(blended.shape[:2], dtype=np.uint8)
    for r in regions:
        cv2.rectangle(mask, (r.x, r.y), (r.x + r.w, r.y + r.h), 255, cv2.FILLED)

    # Red highlight overlay on diff areas
    red = np.zeros_like(blended)
    red[:, :, 2] = 255  # BGR — red channel
    blended[mask > 0] = cv2.addWeighted(blended, 0.5, red, 0.5, 0)[mask > 0]

    # Draw red rectangles around diff regions
    for r in regions:
        cv2.rectangle(blended, (r.x, r.y), (r.x + r.w, r.y + r.h), (0, 0, 255), 2)

    return blended


def count_text_blocks(pdf_path: Path, page_index: int) -> int:
    from pdfminer.high_level import extract_pages
    from pdfminer.layout import LTTextBox

    for i, page_layout in enumerate(extract_pages(str(pdf_path), page_numbers=[page_index])):
        return sum(1 for el in page_layout if isinstance(el, LTTextBox))
    return 0


def decide_status(diff_regions: int, diff_area: int, diff_ratio: float, thresholds: dict) -> Tuple[str, str]:
    if (
        diff_regions >= thresholds["ng_min_regions"]
        or diff_area >= thresholds["ng_min_area"]
        or diff_ratio >= thresholds["ng_min_ratio"]
    ):
        return "NG", "threshold_ng"
    if diff_regions >= thresholds["review_min_regions"] or diff_area >= thresholds["review_min_area"]:
        return "REVIEW", "threshold_review"
    return "OK", "no_diff"


def pair_pdfs(
    old_dir: Path,
    new_dir: Path,
) -> Tuple[List[Tuple[Path, Path]], List[str], List[str], List[str]]:
    old_files = {p.name: p for p in sorted(old_dir.glob("*.pdf"))}
    new_files = {p.name: p for p in sorted(new_dir.glob("*.pdf"))}

    paired = []
    warnings: List[str] = []
    old_only: List[str] = []
    new_only: List[str] = []
    for name in sorted(old_files.keys() & new_files.keys()):
        paired.append((old_files[name], new_files[name]))
    for name in sorted(old_files.keys() - new_files.keys()):
        warnings.append(f"missing in new: {name}")
        old_only.append(name)
    for name in sorted(new_files.keys() - old_files.keys()):
        warnings.append(f"missing in old: {name}")
        new_only.append(name)
    return paired, warnings, old_only, new_only


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def _merge_report_rows(
    results: Sequence[PageResult],
    old_only: Sequence[str],
    new_only: Sequence[str],
) -> List[list]:
    """Merge results and old_only/new_only into a single list sorted by filename."""
    rows: List[Tuple[str, int, list]] = []  # (filename, page, values)
    for r in results:
        rows.append(
            (
                r.file,
                r.page,
                [
                    r.file,
                    r.page,
                    r.old_width,
                    r.old_height,
                    r.new_width,
                    r.new_height,
                    r.status,
                    r.diff_regions,
                    r.diff_area,
                    r.diff_ratio,
                    r.review_reason,
                    r.diff_image or "",
                    r.blocks_old,
                    r.blocks_new,
                    r.block_count_delta,
                ],
            )
        )
    for name in old_only:
        rows.append(
            (name, 0, [name, "", "", "", "", "", "OLD_ONLY", "", "", "", "新ディレクトリに存在しない", "", "", "", ""])
        )
    for name in new_only:
        rows.append(
            (name, 0, [name, "", "", "", "", "", "NEW_ONLY", "", "", "", "旧ディレクトリに存在しない", "", "", "", ""])
        )
    rows.sort(key=lambda x: (x[0], x[1]))
    return [r[2] for r in rows]


_CSV_HEADERS = [
    "ファイル名",
    "ページ",
    "旧_幅",
    "旧_高さ",
    "新_幅",
    "新_高さ",
    "判定",
    "差分領域数",
    "差分面積",
    "差分率",
    "判定理由",
    "差分画像",
    "旧_ブロック数",
    "新_ブロック数",
    "ブロック数差分",
]


def write_csv(
    path: Path,
    results: Sequence[PageResult],
    old_only: Sequence[str],
    new_only: Sequence[str],
) -> None:
    merged = _merge_report_rows(results, old_only, new_only)
    with path.open("w", newline="\n", encoding="utf-8-sig") as f:
        writer = csv.writer(f, lineterminator="\n")
        writer.writerow(_CSV_HEADERS)
        for values in merged:
            csv_values = list(values)
            # Format diff_ratio for CSV
            if isinstance(csv_values[9], float):
                csv_values[9] = f"{csv_values[9]:.8f}"
            writer.writerow(csv_values)


def _scale_image(img: np.ndarray, target_width: int) -> np.ndarray:
    """Resize image to target_width preserving aspect ratio."""
    h, w = img.shape[:2]
    scale = target_width / w
    return cv2.resize(img, (target_width, int(h * scale)), interpolation=cv2.INTER_AREA)


def _sanitize_sheet_name(name: str, max_len: int = 31) -> str:
    """Remove illegal Excel sheet name characters and truncate."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, "_")
    return name[:max_len]


def write_json(path: Path, results: Sequence[PageResult], warnings: Sequence[str], config: dict) -> None:
    payload = {
        "config": config,
        "warnings": list(warnings),
        "results": [asdict(r) for r in results],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


_STATUS_FILLS = {
    "OK": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "REVIEW": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "NG": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "OLD_ONLY": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
    "NEW_ONLY": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
}

_STATUS_COL_INDEX = 7  # 1-based column index for "判定"


def write_xlsx(
    path: Path,
    results: Sequence[PageResult],
    old_only: Sequence[str],
    new_only: Sequence[str],
    page_images: Dict[str, Tuple[Path, Path, Path]],
) -> None:
    wb = openpyxl.Workbook()

    # --- Sheet 1: レポート ---
    ws = wb.active
    ws.title = "レポート"
    header_font = Font(bold=True)
    for col_idx, header in enumerate(_CSV_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    merged = _merge_report_rows(results, old_only, new_only)
    for row_num, values in enumerate(merged, 2):
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_num, column=col_idx, value=val)
        status = values[6]  # "判定" column
        fill = _STATUS_FILLS.get(status)
        if fill:
            ws.cell(row=row_num, column=_STATUS_COL_INDEX).fill = fill

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(_CSV_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # --- Sheet 2+: 画像比較シート ---
    used_names: Dict[str, int] = {}
    for r in results:
        if r.status not in ("REVIEW", "NG"):
            continue
        key = f"{Path(r.file).stem}_p{r.page:03d}"
        if key not in page_images:
            continue
        old_img_path, overlay_img_path, new_img_path = page_images[key]

        sheet_name = _sanitize_sheet_name(key)
        if sheet_name in used_names:
            used_names[sheet_name] += 1
            suffix = f"_{used_names[sheet_name]}"
            sheet_name = _sanitize_sheet_name(key, 31 - len(suffix)) + suffix
        else:
            used_names[sheet_name] = 1

        img_ws = wb.create_sheet(title=sheet_name)

        img_ws.cell(row=1, column=1, value="旧 (OLD)").font = Font(bold=True)
        img_ws.cell(row=1, column=2, value="差分 (OVERLAY)").font = Font(bold=True)
        img_ws.cell(row=1, column=3, value="新 (NEW)").font = Font(bold=True)
        img_ws.cell(row=2, column=1, value=f"{r.file}  p.{r.page}  [{r.status}]  差分率: {r.diff_ratio:.6f}")
        img_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
        img_ws.cell(row=2, column=1).alignment = Alignment(horizontal="left")

        col_width_chars = 80
        for col in range(1, 4):
            img_ws.column_dimensions[get_column_letter(col)].width = col_width_chars

        for col_idx, img_path in enumerate([old_img_path, overlay_img_path, new_img_path], 1):
            if img_path.exists():
                img = XlImage(str(img_path))
                anchor = f"{get_column_letter(col_idx)}3"
                img_ws.add_image(img, anchor)

        img_height_px = 792  # approx A4 at 560px width
        row_height_pts = img_height_px * 0.75
        img_ws.row_dimensions[3].height = row_height_pts

    wb.save(str(path))


def compare_pair(
    old_pdf: Path,
    new_pdf: Path,
    config: dict,
    output_dir: Path,
    save_diff_image: bool = True,
    temp_dir: Optional[Path] = None,
) -> Tuple[List[PageResult], Dict[str, Tuple[Path, Path, Path]]]:
    """Compare a pair of PDFs page-by-page.

    Returns (page_results, page_images) where page_images maps
    ``"{stem}_p{page:03d}"`` to ``(old_png, overlay_png, new_png)`` paths
    saved under *temp_dir* (only for REVIEW/NG pages when temp_dir is set).
    """
    old_pages = render_pdf_pages(old_pdf, config["dpi"])
    new_pages = render_pdf_pages(new_pdf, config["dpi"])

    if len(old_pages) != len(new_pages):
        policy = config["page_count_policy"]
        if policy == "error":
            raise ValueError(f"page count mismatch: {old_pdf.name} old={len(old_pages)} new={len(new_pages)}")
        count = min(len(old_pages), len(new_pages))
    else:
        count = len(old_pages)

    if save_diff_image:
        ensure_dir(output_dir)

    page_results: List[PageResult] = []
    page_images: Dict[str, Tuple[Path, Path, Path]] = {}
    for page_index in range(count):
        img_old, img_new = normalize_size(old_pages[page_index], new_pages[page_index], config["crop_to_common_size"])

        if config["align_images"]:
            img_new = try_align_images(
                img_old,
                img_new,
                iterations=int(config["ecc_iterations"]),
                eps=float(config["ecc_eps"]),
            )

        _, regions, diff_area, diff_ratio = detect_diff(
            img_old,
            img_new,
            pixel_threshold=int(config["pixel_threshold"]),
            min_region_area=int(config["min_region_area"]),
            morph_kernel=int(config["morph_kernel"]),
        )
        status, reason = decide_status(len(regions), diff_area, diff_ratio, config["status_thresholds"])

        diff_image_path = None
        if status in ("REVIEW", "NG"):
            need_overlay = save_diff_image or temp_dir is not None
            overlay = create_overlay(img_old, img_new, regions) if need_overlay else None

            if save_diff_image and overlay is not None:
                diff_image_path = output_dir / f"{old_pdf.stem}_p{page_index + 1:03d}_diff.png"
                imwrite_unicode(diff_image_path, overlay)

            if temp_dir is not None and overlay is not None:
                ensure_dir(temp_dir)
                key = f"{old_pdf.stem}_p{page_index + 1:03d}"
                old_path = temp_dir / f"{key}_old.png"
                overlay_path = temp_dir / f"{key}_overlay.png"
                new_path = temp_dir / f"{key}_new.png"
                imwrite_unicode(old_path, _scale_image(img_old, 560))
                imwrite_unicode(overlay_path, _scale_image(overlay, 560))
                imwrite_unicode(new_path, _scale_image(img_new, 560))
                page_images[key] = (old_path, overlay_path, new_path)

        blocks_old = count_text_blocks(old_pdf, page_index)
        blocks_new = count_text_blocks(new_pdf, page_index)

        page_results.append(
            PageResult(
                file=old_pdf.name,
                page=page_index + 1,
                old_width=int(img_old.shape[1]),
                old_height=int(img_old.shape[0]),
                new_width=int(img_new.shape[1]),
                new_height=int(img_new.shape[0]),
                status=status,
                diff_regions=len(regions),
                diff_area=diff_area,
                diff_ratio=diff_ratio,
                review_reason=reason,
                diff_image=str(diff_image_path.relative_to(output_dir)) if diff_image_path else None,
                blocks_old=blocks_old,
                blocks_new=blocks_new,
                block_count_delta=blocks_new - blocks_old,
                regions=regions,
            )
        )
    return page_results, page_images


def write_with_retry(path: Path, write_fn) -> None:
    """Try writing a file, prompting user to retry on PermissionError."""
    while True:
        try:
            write_fn(path)
            return
        except PermissionError:
            print(f"\nERROR: ファイルを書き込めません: {path}", file=sys.stderr)
            print("  → Excelなどで開いている場合は閉じてください。", file=sys.stderr)
            answer = input("  リトライしますか？ (Y/n): ").strip().lower()
            if answer in ("n", "no"):
                print(f"  スキップしました: {path}", file=sys.stderr)
                return


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compare layout differences between paired PDFs.")
    parser.add_argument("--version", action="version", version=f"%(prog)s {__version__}")
    parser.add_argument("--old-dir", type=Path, required=True, help="Directory containing previous PDFs")
    parser.add_argument("--new-dir", type=Path, required=True, help="Directory containing current PDFs")
    parser.add_argument("--output-dir", type=Path, default=Path("output"), help="Directory for reports and diff images")
    parser.add_argument("--config", type=Path, default=Path("config.yaml"), help="YAML config path")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N paired PDFs")
    parser.add_argument(
        "--format", choices=["csv", "json", "xlsx"], default="xlsx", help="Output report format (default: xlsx)"
    )
    parser.add_argument(
        "--without-diff-image",
        action="store_true",
        default=False,
        help="Skip diff image file generation (diff_images/ directory)",
    )
    return parser


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if not args.old_dir.is_dir():
        parser.error(f"old-dir not found: {args.old_dir}")
    if not args.new_dir.is_dir():
        parser.error(f"new-dir not found: {args.new_dir}")

    config_path = args.config if args.config.exists() else None
    config = load_config(config_path)

    ensure_dir(args.output_dir)

    fmt = args.format
    save_diff_image = not args.without_diff_image
    need_image_sheets = fmt == "xlsx" and save_diff_image

    # Build list of targets to check for overwrite
    report_ext = {"csv": "report.csv", "json": "report.json", "xlsx": "report.xlsx"}
    targets = [args.output_dir / report_ext[fmt]]
    if save_diff_image:
        targets.append(args.output_dir / "diff_images")

    existing = [t for t in targets if t.exists()]
    if existing:
        print("以下の出力先が既に存在します:")
        for p in existing:
            print(f"  {p}")
        answer = input("上書き（削除して再作成）しますか？ (Y/n): ").strip().lower()
        if answer in ("n", "no"):
            print("中断しました。")
            return 1
        for p in existing:
            if p.is_dir():
                shutil.rmtree(p)
            else:
                p.unlink()
        print("既存の出力を削除しました。\n")

    paired, warnings, old_only, new_only = pair_pdfs(args.old_dir, args.new_dir)
    if args.limit is not None:
        paired = paired[: args.limit]

    if not paired:
        print("No matching PDF filenames found.", file=sys.stderr)
        return 2

    temp_dir = args.output_dir / ".temp_images" if need_image_sheets else None

    all_results: List[PageResult] = []
    all_page_images: Dict[str, Tuple[Path, Path, Path]] = {}
    try:
        for old_pdf, new_pdf in paired:
            try:
                results, images = compare_pair(
                    old_pdf,
                    new_pdf,
                    config,
                    args.output_dir / "diff_images",
                    save_diff_image=save_diff_image,
                    temp_dir=temp_dir,
                )
                all_results.extend(results)
                all_page_images.update(images)
                print(f"OK: {old_pdf.name} ({len(results)} pages)")
            except Exception as exc:
                warnings.append(f"{old_pdf.name}: {exc}")
                print(f"ERROR: {old_pdf.name}: {exc}", file=sys.stderr)

        if fmt == "csv":
            write_with_retry(args.output_dir / "report.csv", lambda p: write_csv(p, all_results, old_only, new_only))
        elif fmt == "json":
            write_with_retry(args.output_dir / "report.json", lambda p: write_json(p, all_results, warnings, config))
        elif fmt == "xlsx":
            write_with_retry(
                args.output_dir / "report.xlsx",
                lambda p: write_xlsx(p, all_results, old_only, new_only, all_page_images),
            )
    finally:
        if temp_dir is not None and temp_dir.exists():
            shutil.rmtree(temp_dir)

    summary = {
        "files": len({r.file for r in all_results}),
        "pages": len(all_results),
        "ok": sum(1 for r in all_results if r.status == "OK"),
        "review": sum(1 for r in all_results if r.status == "REVIEW"),
        "ng": sum(1 for r in all_results if r.status == "NG"),
        "warnings": len(warnings),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if summary["warnings"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
